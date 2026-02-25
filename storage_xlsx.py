"""
Excel persistence for Daily Task Board.

Workbook format (required by your spec):
- Sheets: Tasks, Finished, Unfinished, Working, Postponned
- Every row stores Start date+time, and Finished date+time if applicable.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.colors import Color

from models import DailyTask, TaskStatus, ProjectItem

SHEETS = ["Tasks", "Finished", "Unfinished", "Working", "Postponned"]
# NOTE: schema is intentionally append-only and backwards compatible.
HEADERS = [
    "ID",
    "Title",
    "Status",
    "Start",
    "Due",
    "Finished",
    "Order",
    "DetailsHtml",
    "IsNote",
    "Attachments",
    "RemindAt",
    "Remind10m",
    "Remind1h",
    "SentAt",
    "Sent10m",
    "Sent1h",
]

PROJECT_SHEET = "Projects"
PROJECT_HEADERS = [
    "ID",
    "Title",
    "Status",
    "Category",
    "CodePath",
    "Links",
    "Reminder",
    "Updated",
    "Finished",
    "Order",
    "DetailsHtml",
    "Attachments",
    "RemindAt",
    "Remind10m",
    "Remind1h",
    "SentAt",
    "Sent10m",
    "Sent1h",
]

DT_FORMAT = "yyyy-mm-dd hh:mm"

# --- Styling (Excel) ---
# NOTE: openpyxl expects ARGB colors (8 hex digits). We use FF + RGB.
FILL_HEADER = PatternFill(fill_type="solid", fgColor=Color(rgb="FFF3F4F6"))  # light gray
FONT_HEADER = Font(bold=True)
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FILL_STATUS = {
    TaskStatus.UNFINISHED.value: PatternFill(fill_type="solid", fgColor=Color(rgb="FFF4CCCC")),  # light red
    TaskStatus.WORKING.value: PatternFill(fill_type="solid", fgColor=Color(rgb="FFD9EAF7")),     # light blue
    TaskStatus.POSTPONED.value: PatternFill(fill_type="solid", fgColor=Color(rgb="FFFFF2CC")),   # light amber
    TaskStatus.FINISHED.value: PatternFill(fill_type="solid", fgColor=Color(rgb="FFD9EAD3")),    # light green
}
FILL_NOTE = PatternFill(fill_type="solid", fgColor=Color(rgb="FFF2F2F2"))  # neutral gray

# Sheet/tab colors (ARGB)
TAB_COLORS = {
    "Tasks": "FF9FC5E8",        # soft blue
    "Unfinished": "FFF4CCCC",   # soft red
    "Working": "FFD9EAF7",      # soft blue
    "Postponned": "FFFFF2CC",   # soft amber (typo kept for backwards compatibility)
    "Postponed": "FFFFF2CC",
    "Finished": "FFD9EAD3",     # soft green
    "Projects": "FFD9D2E9",     # soft purple
}


def _style_header_row(ws, header_count: int) -> None:
    """Apply header styling."""
    for c in range(1, header_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_HEADER
    ws.row_dimensions[1].height = 22


def _apply_row_fill(ws, status_col: int, is_note_col: Optional[int] = None) -> None:
    """Color rows by status (and notes if applicable)."""
    for r in range(2, ws.max_row + 1):
        status = ws.cell(row=r, column=status_col).value
        is_note = False
        if is_note_col is not None:
            v = ws.cell(row=r, column=is_note_col).value
            is_note = bool(v) if v is not None else False

        fill = FILL_NOTE if is_note else FILL_STATUS.get(str(status), None)
        if fill is None:
            continue

        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill


def _set_tab_color(ws, name: str) -> None:
    rgb = TAB_COLORS.get(name)
    if not rgb:
        return
    try:
        ws.sheet_properties.tabColor = Color(rgb=rgb)
    except Exception:
        # fallback if older openpyxl expects raw string
        try:
            ws.sheet_properties.tabColor = rgb
        except Exception:
            pass


def _autosize(ws):
    # Basic autosize based on header / first ~100 rows
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for r in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 101), min_col=col_idx, max_col=col_idx):
            v = r[0].value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)


def _write_sheet(wb: Workbook, name: str, tasks: List[DailyTask]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
        wb.remove(ws)
    ws = wb.create_sheet(name)
    ws.append(HEADERS)
    ws.freeze_panes = "A2"

    _set_tab_color(ws, name)
    _style_header_row(ws, len(HEADERS))

    for t in tasks:
        ws.append(t.to_row())

    # Date formatting: Start(4), Due(5), Finished(6)
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = DT_FORMAT

    # Color rows by status / notes
    # Columns: Status = 3, IsNote = 9
    _apply_row_fill(ws, status_col=3, is_note_col=9)

    _autosize(ws)



def _write_projects_sheet(wb: Workbook, projects: List[ProjectItem]) -> None:
    if PROJECT_SHEET in wb.sheetnames:
        ws = wb[PROJECT_SHEET]
        wb.remove(ws)
    ws = wb.create_sheet(PROJECT_SHEET)
    ws.append(PROJECT_HEADERS)
    ws.freeze_panes = "A2"

    _set_tab_color(ws, PROJECT_SHEET)
    _style_header_row(ws, len(PROJECT_HEADERS))

    for p in projects:
        ws.append(p.to_row())

    # Date formatting: Reminder(7), Updated(8), Finished(9)
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=9):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = DT_FORMAT

    # Color rows by status (Status = 3)
    _apply_row_fill(ws, status_col=3, is_note_col=None)

    # Autosize based on project headers
    for col_idx, header in enumerate(PROJECT_HEADERS, start=1):
        max_len = len(header)
        for r in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 101), min_col=col_idx, max_col=col_idx):
            v = r[0].value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

def save_data_xlsx(path: str, tasks: List[DailyTask], projects: Optional[List[ProjectItem]] = None) -> None:
    """Save tasks + projects into a single Excel workbook."""
    projects = projects or []
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    all_tasks = list(tasks)

    unfinished = [t for t in all_tasks if t.status == TaskStatus.UNFINISHED]
    finished = [t for t in all_tasks if t.status == TaskStatus.FINISHED]
    working = [t for t in all_tasks if t.status == TaskStatus.WORKING]
    postponed = [t for t in all_tasks if t.status == TaskStatus.POSTPONED]

    _write_sheet(wb, "Tasks", all_tasks)
    _write_sheet(wb, "Finished", finished)
    _write_sheet(wb, "Unfinished", unfinished)
    _write_sheet(wb, "Working", working)
    _write_sheet(wb, "Postponned", postponed)

    _write_projects_sheet(wb, projects)

    wb.save(str(p))


def save_tasks_xlsx(path: str, tasks: List[DailyTask]) -> None:
    """Backwards-compatible wrapper (tasks only)."""
    save_data_xlsx(path, tasks, projects=[])




def _read_sheet_rows(wb, name: str) -> List[Dict]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None:
            continue
        row_dict = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            if v not in (None, ""):
                empty = False
            row_dict[h] = v
        if empty:
            continue
        out.append(row_dict)
    return out


def load_data_xlsx(path: str):
    """Load tasks + projects from an Excel workbook.

    Returns: (tasks, projects)
    """
    p = Path(path)
    if not p.exists():
        return [], []

    wb = openpyxl.load_workbook(str(p))

    # --- Tasks ---
    rows = _read_sheet_rows(wb, "Tasks")
    if not rows:
        merged = []
        for n in ("Unfinished", "Working", "Finished", "Postponned", "Postponed"):
            merged.extend(_read_sheet_rows(wb, n))
        rows = merged

    tasks: List[DailyTask] = []
    seen = set()
    for rd in rows:
        t = DailyTask.from_row(rd)
        if not t.id:
            continue
        if t.id in seen:
            continue
        seen.add(t.id)
        tasks.append(t)

    # --- Projects ---
    proj_rows = _read_sheet_rows(wb, PROJECT_SHEET)
    projects: List[ProjectItem] = []
    pseen = set()
    for rd in proj_rows:
        pr = ProjectItem.from_row(rd)
        if not pr.id:
            continue
        if pr.id in pseen:
            continue
        pseen.add(pr.id)
        projects.append(pr)

    return tasks, projects


def load_tasks_xlsx(path: str) -> List[DailyTask]:
    """Backwards-compatible wrapper (tasks only)."""
    tasks, _ = load_data_xlsx(path)
    return tasks
